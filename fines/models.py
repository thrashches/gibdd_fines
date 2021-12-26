from django.db import models, transaction
from django.db.utils import IntegrityError
from django.contrib.auth.models import AbstractUser
from openpyxl import load_workbook
from datetime import datetime


# Create your models here.
class Driver(models.Model):
    class Meta:
        verbose_name = 'водитель'
        verbose_name_plural = 'водители'

    full_name = models.CharField(max_length=255, verbose_name='ФИО')

    def __str__(self):
        return f'{self.full_name}'


class Fine(models.Model):
    class Meta:
        verbose_name = 'штраф'
        verbose_name_plural = 'штрафы'
        ordering = ['-fine_date']

    accident_number = models.CharField(max_length=25, unique=True, verbose_name='номер постановления')
    price = models.IntegerField(blank=True, null=True, verbose_name='сумма штрафа')
    accident_date = models.DateField(blank=True, null=True, verbose_name='дата нарушения')
    accident_time = models.TimeField(blank=True, null=True, verbose_name='время нарушения')
    accident_datetime = models.DateTimeField(blank=True, null=True, verbose_name='дата и время нарушения')
    fine_date = models.DateField(blank=True, null=True, verbose_name='дата вынесения постановления')
    car = models.ForeignKey('Car', related_name='fines', null=True, blank=True, on_delete=models.CASCADE,
                            verbose_name='автомобиль')
    pay_status = models.BooleanField(default=False, verbose_name='статус оплаты')
    driver = models.ForeignKey(Driver, blank=True, null=True, on_delete=models.CASCADE, verbose_name='водитель')

    def __str__(self):
        return f'{self.accident_number}'

    def save(self, force_insert=False, force_update=False, using=None,
             update_fields=None):
        if self.accident_time and self.accident_date:
            self.accident_datetime = datetime.combine(self.accident_date, self.accident_time)
        super(Fine, self).save()


class Car(models.Model):
    class Meta:
        verbose_name = 'автомобиль'
        verbose_name_plural = 'автомобили'

    auto_id = models.IntegerField(unique=True, verbose_name='id автомобиля в API')
    auto_number = models.CharField(max_length=10, blank=True, null=True, verbose_name='гос номер(без региона)')
    auto_region = models.CharField(max_length=3, blank=True, null=True, verbose_name='регион')
    auto_cdi = models.CharField(max_length=10, blank=True, null=True, verbose_name='номер стс')
    auto_name = models.CharField(max_length=255, blank=True, null=True, verbose_name='название автомобиля')

    def __str__(self):
        return f'{self.auto_id}: {self.auto_number}{self.auto_region}'


class Profile(AbstractUser):
    pass


class Report(models.Model):
    class Meta:
        verbose_name = 'отчет'
        verbose_name_plural = 'отчеты'

    file = models.FileField(upload_to='media', verbose_name='excel файл')
    upload_date = models.DateTimeField(auto_now_add=True)

    def __str__(self):
        return str(self.upload_date)

    def save(self, force_insert=False, force_update=False, using=None,
             update_fields=None):
        if not self.id:
            wb = load_workbook(filename=self.file.path)
            ws = wb.worksheets[0]
            # drivers_data = []
            line = 1
            while True:
                line_data = {
                    'begin_time': ws['A' + str(line)].value,
                    'end_time': ws['B' + str(line)].value,
                    'driver_name': ws['C' + str(line)].value,
                    'car_number': ws['D' + str(line)].value
                }
                print('line_data:', line_data)
                if not line_data['begin_time']:
                    break
                else:
                    line += 1
                    driver = Driver.objects.filter(full_name=line_data['driver_name']).first()
                    car = Car.objects.filter(auto_number__icontains=str(line_data['car_number'])).first()
                    try:
                        with transaction.atomic():
                            driver_on_car = DriverOnCar.objects.create(
                                begin_time=line_data['begin_time'],
                                end_time=line_data['end_time'],
                                car=car,
                                driver=driver
                            )
                    except IntegrityError:
                        continue

        super(Report, self).save()


class DriverOnCar(models.Model):
    class Meta:
        verbose_name = 'учет времени за рулем'
        verbose_name_plural = 'учет времени за рулем'
        unique_together = ['driver', 'car', 'begin_time', 'end_time']

    driver = models.ForeignKey(Driver, on_delete=models.CASCADE, verbose_name='водитель')
    car = models.ForeignKey(Car, on_delete=models.CASCADE, verbose_name='автомобиль')
    begin_time = models.DateTimeField(verbose_name='время получения авто')
    end_time = models.DateTimeField(verbose_name='время сдачи авто')

    def __str__(self):
        return f'{self.car.auto_number}: {self.begin_time} - {self.end_time}'

    def save(self, force_insert=False, force_update=False, using=None,
             update_fields=None):
        if not self.id:
            fines = Fine.objects.filter(accident_datetime__range=(self.begin_time, self.end_time)).filter(car=self.car)
            print(fines)
            fines.update(driver=self.driver)

        super(DriverOnCar, self).save()
